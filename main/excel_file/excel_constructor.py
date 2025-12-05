import os
from io import BytesIO
from django.http import HttpResponse
from django.conf import settings
from openpyxl import load_workbook
from openpyxl.styles import Alignment

MESES_ORDEM = [
    "janeiro", "fevereiro", "marco", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
]


def create_projetos_excel_response(projetos, filename="PLANEJAMENTO ORÇAMENTÁRIO.xlsx", gestor=None, setor=None, centro_custo=None):

    template_path = os.path.join(
        settings.BASE_DIR,
        "main",
        "excel_file",
        "modelo_projetos.xlsx"
    )

    wb = load_workbook(template_path)

    ws_proj = wb["Projetos"]
    ws_orc = wb["Orçamento"]

    # ======================================================================
    # ABA PROJETOS
    # ======================================================================
    LINHA_PROJ = 5  # onde começam os projetos

    # Preencher Gestor e Setor (se fornecidos)
    if gestor:
        ws_proj["B2"] = gestor
        ws_proj["B2"].alignment = Alignment(wrap_text=True)
    if setor:
        ws_proj["D2"] = setor
        ws_proj["D2"].alignment = Alignment(wrap_text=True)
    # escreve Centro de Custo na aba Orçamento célula A2 se informado
    if centro_custo:
        try:
            ws_orc["A2"] = centro_custo
            ws_orc["A2"].alignment = Alignment(wrap_text=True)
        except Exception:
            pass

    linha_p = LINHA_PROJ

    for p in projetos:

        valores = p.valores_mensais or {}

        # Nome do projeto
        ws_proj[f"A{linha_p}"] = p.nome_projeto

        # Objetivo / Justificativa → escreve em B5, B6, ...
        justificativa = getattr(p, 'justificativa', '') or ''
        ws_proj[f"B{linha_p}"] = justificativa
        # aplica quebra de linha e alinhamento para justificativa
        cell_just = ws_proj[f"B{linha_p}"]
        cell_just.alignment = Alignment(wrap_text=True, vertical='top')

        # Recorrência
        ws_proj[f"C{linha_p}"] = "Sim" if p.e_recorrente else "Não"

        # Obrigação Legal
        ws_proj[f"D{linha_p}"] = "Sim" if p.obrigacao_legal else "Não"

        # Valor total
        ws_proj[f"E{linha_p}"] = p.valor_total

        linha_p += 1

    # ======================================================================
    # ABA ORÇAMENTO
    # ======================================================================
    LINHA_ORC = 4

    linha_o = LINHA_ORC

    for p in projetos:
        # Se existirem contas vinculadas, escrever uma linha por conta
        contas = getattr(p, 'contas_valores', None)
        if contas is not None and hasattr(contas, 'all') and contas.all():
            for pcv in contas.all():
                ws_orc[f"A{linha_o}"] = p.nome_projeto
                ws_orc[f"B{linha_o}"] = p.tipo_conta
                ws_orc[f"C{linha_o}"] = pcv.conta_contabil
                col_base = 4
                for i, mes in enumerate(MESES_ORDEM):
                    col = col_base + i
                    val = (pcv.valores_mensais or {}).get(mes)
                    ws_orc.cell(row=linha_o, column=col).value = float(val) if val else 0
                ws_orc[f"P{linha_o}"] = pcv.valor_total
                linha_o += 1
        else:
            # fallback: projeto tem valores mensais no nível do projeto
            valores = p.valores_mensais or {}
            ws_orc[f"A{linha_o}"] = p.nome_projeto
            ws_orc[f"B{linha_o}"] = p.tipo_conta
            ws_orc[f"C{linha_o}"] = p.codigo_produto
            col_base = 4
            for i, mes in enumerate(MESES_ORDEM):
                col = col_base + i
                val = valores.get(mes)
                ws_orc.cell(row=linha_o, column=col).value = float(val) if val else 0
            ws_orc[f"P{linha_o}"] = p.valor_total
            linha_o += 1

    # ======================================================================
    # RETORNO DO ARQUIVO GERADO
    # ======================================================================
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    response = HttpResponse(
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response
